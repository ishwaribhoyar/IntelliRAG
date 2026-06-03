"""Document content extractors — PyMuPDF, Docling fallback, OCR, Excel.
Upgraded: text cleaning (header/footer removal, line-break fix, whitespace normalization),
page-number tracking per section.
"""
import re
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Text cleaning helpers
# ---------------------------------------------------------------------------

def _clean_text(text: str) -> str:
    """Clean raw extracted text: fix line-breaks, normalise whitespace."""
    if not text:
        return ""
    # Fix mid-word line wraps (word-\nword → word-word, word\nword → word word)
    text = re.sub(r'(\w)-\n(\w)', r'\1\2', text)
    text = re.sub(r'(?<!\n)\n(?!\n)', ' ', text)  # single newline → space
    # Collapse multiple spaces
    text = re.sub(r'[ \t]+', ' ', text)
    # Collapse 3+ newlines into 2
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _remove_headers_footers(pages_text: list[str]) -> list[str]:
    """Detect repeated first/last lines across pages and remove them.

    Heuristic: if the same line appears as the first (or last) line on ≥50%
    of pages, it is a header (or footer).
    """
    if len(pages_text) < 3:
        return pages_text  # too few pages to detect pattern

    # Collect first and last non-empty lines
    first_lines: dict[str, int] = {}
    last_lines: dict[str, int] = {}
    for page in pages_text:
        lines = [l.strip() for l in page.split('\n') if l.strip()]
        if lines:
            first_lines[lines[0]] = first_lines.get(lines[0], 0) + 1
            last_lines[lines[-1]] = last_lines.get(lines[-1], 0) + 1

    threshold = len(pages_text) * 0.5
    header_lines = {l for l, c in first_lines.items() if c >= threshold and len(l) < 120}
    footer_lines = {l for l, c in last_lines.items() if c >= threshold and len(l) < 120}

    if not header_lines and not footer_lines:
        return pages_text

    cleaned = []
    for page in pages_text:
        lines = page.split('\n')
        out = []
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped in header_lines and i < 3:
                continue
            if stripped in footer_lines and i > len(lines) - 4:
                continue
            out.append(line)
        cleaned.append('\n'.join(out))

    removed = len(header_lines) + len(footer_lines)
    if removed:
        logger.info(f"Removed {len(header_lines)} header patterns, {len(footer_lines)} footer patterns")
    return cleaned


# ---------------------------------------------------------------------------
# Extractors
# ---------------------------------------------------------------------------

def extract_with_pymupdf(file_path: str) -> dict:
    """Fast extraction using PyMuPDF / pymupdf4llm with per-page text + cleaning."""
    import pymupdf4llm
    import pymupdf

    doc = pymupdf.open(file_path)
    try:
        page_count = len(doc)

        # Per-page raw text (for header/footer detection)
        pages_raw: list[str] = []
        for page in doc:
            pages_raw.append(page.get_text("text"))

        # Remove headers/footers
        pages_cleaned = _remove_headers_footers(pages_raw)

        # Also get the markdown version for structure
        md_text = pymupdf4llm.to_markdown(doc)
        md_text = _clean_text(md_text)
    finally:
        doc.close()

    return {
        "type": "markdown",
        "content": md_text,
        "page_count": page_count,
        "pages_text": [_clean_text(p) for p in pages_cleaned],
    }


def extract_with_docling(file_path: str) -> dict:
    """Structured extraction using Docling for complex layouts."""
    try:
        from docling.document_converter import DocumentConverter
        converter = DocumentConverter()
        result = converter.convert(file_path)
        text = result.document.export_to_markdown()
        text = _clean_text(text)
        return {
            "type": "structured",
            "content": text,
            "page_count": 1,
            "pages_text": [text],
        }
    except Exception as e:
        logger.warning(f"Docling failed, falling back to PyMuPDF: {e}")
        return extract_with_pymupdf(file_path)


def extract_with_ocr(file_path: str) -> dict:
    """OCR extraction using Tesseract via PyMuPDF."""
    import pymupdf

    doc = pymupdf.open(file_path)
    try:
        pages_text: list[str] = []
        for page in doc:
            text = page.get_text("text")
            if not text.strip():
                try:
                    tp = page.get_textpage_ocr(language="eng", full=True)
                    text = page.get_text("text", textpage=tp)
                except Exception:
                    text = ""
            pages_text.append(text)

        page_count = len(doc)
    finally:
        doc.close()

    pages_cleaned = _remove_headers_footers(pages_text)
    pages_cleaned = [_clean_text(p) for p in pages_cleaned]
    full_text = "\n\n".join(pages_cleaned)

    return {
        "type": "ocr",
        "content": full_text,
        "page_count": page_count,
        "pages_text": pages_cleaned,
    }


def extract_excel(file_path: str) -> dict:
    """Extract content from Excel files."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            if any(v.strip() for v in row_values):
                rows.append(" | ".join(row_values))

        content = "\n".join(rows)
        sections.append({"heading": sheet_name, "content": content})

    wb.close()

    combined = "\n\n".join(f"## {s['heading']}\n{s['content']}" for s in sections)

    return {
        "type": "excel",
        "content": combined,
        "page_count": len(sections),
        "pages_text": [s["content"] for s in sections],
        "sections": sections,
    }


def extract_document(file_path: str, parser_type: str) -> dict:
    """Main extraction dispatcher with fallback chain."""
    ext = Path(file_path).suffix.lower()

    if ext == ".xlsx":
        return extract_excel(file_path)

    try:
        if parser_type == "ocr":
            return extract_with_ocr(file_path)
        elif parser_type == "docling":
            return extract_with_docling(file_path)
        else:
            return extract_with_pymupdf(file_path)
    except Exception as e:
        logger.error(f"Primary extraction ({parser_type}) failed: {e}")
        # Fallback chain
        if parser_type == "pymupdf":
            logger.info("Falling back to Docling...")
            try:
                return extract_with_docling(file_path)
            except Exception as e2:
                logger.error(f"Docling fallback also failed: {e2}")
                raise
        elif parser_type == "docling":
            logger.info("Falling back to PyMuPDF...")
            return extract_with_pymupdf(file_path)
        else:
            raise
"""Upgraded extractors with text cleaning and page tracking."""
